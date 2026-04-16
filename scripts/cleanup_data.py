#!/usr/bin/env python3
"""
cleanup_data.py — BankInfoHub One-Shot Data Quality Fix
=========================================================
Fixes ALL data quality issues in your Neon PostgreSQL database:

  1. Districts — removes junk entries (Urban, Rural, Metro, Semi Urban etc),
                 merges spelling variants, removes non-district localities
  2. Pincodes  — extracts 6-digit pincodes from ADDRESS field for all 177K branches
  3. Bank metadata — adds short_name, bank_type, website for major Indian banks
  4. State codes — adds ISO 3166-2 codes for all Indian states

Run from your project root:
    python scripts/cleanup_data.py

Requirements: pip install psycopg2-binary tqdm colorama
"""

import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

# ── Auto-install deps ─────────────────────────────────────────────────────────
try:
    import psycopg2
    import psycopg2.extras
    from tqdm import tqdm
    from colorama import Fore, Style, init as colorama_init
    import openpyxl
except ImportError:
    print("📦 Installing dependencies...")
    os.system(f"{sys.executable} -m pip install psycopg2-binary tqdm colorama openpyxl -q")
    import psycopg2
    import psycopg2.extras
    from tqdm import tqdm
    from colorama import Fore, Style, init as colorama_init
    import openpyxl

colorama_init(autoreset=True)

# ── Load .env ─────────────────────────────────────────────────────────────────
SCRIPTS_DIR = Path(__file__).parent.resolve()
BASE_DIR    = SCRIPTS_DIR.parent

for candidate in [BASE_DIR / ".env", SCRIPTS_DIR / ".env"]:
    if candidate.exists():
        with open(candidate) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k and not os.environ.get(k):
                    os.environ[k] = v
        print(f"✅ Loaded env from: {candidate}")
        break


# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURATION DATA
# ═════════════════════════════════════════════════════════════════════════════

# ── Defunct/Closed/Merged Banks ───────────────────────────────────────────────
# These get deleted by remove_defunct_banks() as a fallback safety net.
# Primary handling is via Bank.xlsx (is_active=false, merged_into_id).
DEFUNCT_BANKS = {
    # 2019-2020 PSB mega-mergers
    "Andhra Bank",                    # → Union Bank of India (2020)
    "Vijaya Bank",                    # → Bank of Baroda (2019)
    "Dena Bank",                      # → Bank of Baroda (2019)
    "Allahabad Bank",                 # → Indian Bank (2020)
    "Syndicate Bank",                 # → Canara Bank (2020)
    "Corporation Bank",               # → Union Bank of India (2020)
    "Oriental Bank of Commerce",      # → Punjab National Bank (2020)
    "United Bank of India",           # → Punjab National Bank (2020)
    # Private bank mergers
    "Laxmi Vilas Bank",               # → DBS Bank India (2020)
    # Junk/test entries
    "Unknown",
}

# ── Junk district names to delete entirely ────────────────────────────────────
# These are RBI branch category codes / generic words — not real districts
JUNK_DISTRICTS = {
    "Urban", "Rural", "Metro", "Semi Urban", "Semiurban", "Semi-Urban",
    "Urban Bank", "Rural Bank", "Metropolitan", "Suburban",
    "Unknown", "Na", "N/A", "Nil", "None", "Test", "Dummy", "Other",
    # Localities that are not revenue districts
    "Kukatpally", "Malkajgiri", "Ameerpet", "Secunderabad",
    "Hanamkonda", "Kodad", "Godavarikhani", "Serilingampally",
    "Seri Lingampally", "Seri-Lingampally",
    "Jammikunta", "Armoor", "Bhongir", "Bodhan", "Huzurabad",
    "Gadwal", "Sircilla", "Bhongir",
}

# ── Canonical district name mapping ──────────────────────────────────────────
# Maps bad/variant names → correct official name
# Format: "bad name as stored in DB (Title Case)" : "Correct Name"
DISTRICT_CANONICAL = {
    # Telangana
    "Rangareddi":               "Rangareddy",
    "Ranga Reddy":              "Rangareddy",
    "Hyderabad-I":              "Hyderabad",
    "Hyderabad-Ii":             "Hyderabad",
    "Hyderabad - I":            "Hyderabad",
    "Hyderabad - Ii":           "Hyderabad",
    "Mahabubnagar":             "Mahbubnagar",
    "Mahabubabad":              "Mahbubnagar",
    "Kothagudem":               "Bhadradri Kothagudem",
    "Warangal Urban":           "Warangal",
    "Warangal Rural":           "Warangal",
    "Secunderabad Cantonment":  "Hyderabad",
    "Yadadri":                  "Yadadri Bhuvanagiri",
    "Jogulamba Gadwal":         "Jogulamba",

    # Andhra Pradesh
    "Visakhapatnam":            "Visakhapatnam",
    "Vizag":                    "Visakhapatnam",
    "Vishakhapatnam":           "Visakhapatnam",
    "Sri Potti Sriramulu Nellore": "Nellore",
    "Spsr Nellore":             "Nellore",

    # Karnataka
    "Bangalore":                "Bengaluru",
    "Bangalore Rural":          "Bengaluru Rural",
    "Bangalore Urban":          "Bengaluru Urban",
    "Bangaluru":                "Bengaluru",

    # Maharashtra
    "Mum Bai":                  "Mumbai",
    "Mumbai Suburban":          "Mumbai",
    "Mumbai City":              "Mumbai",
    "Thane District":           "Thane",

    # Delhi
    "New Delhi":                "Delhi",
    "Newdelhi":                 "Delhi",
    "New Delhi District":       "Delhi",
    "North Delhi":              "Delhi",
    "South Delhi":              "Delhi",
    "East Delhi":               "Delhi",
    "West Delhi":               "Delhi",
    "Central Delhi":            "Delhi",
    "North East Delhi":         "Delhi",
    "North West Delhi":         "Delhi",
    "South East Delhi":         "Delhi",
    "South West Delhi":         "Delhi",
    "Shahdara":                 "Delhi",

    # Punjab
    "S A S Nagar Mohali":       "Sahibzada Ajit Singh Nagar",
    "S.A.S. Nagar Mohali":      "Sahibzada Ajit Singh Nagar",
    "S.A.S.Nagar Mohali":       "Sahibzada Ajit Singh Nagar",
    "Sa S Nagar Mohali":        "Sahibzada Ajit Singh Nagar",
    "Sas Nagar - Mohali":       "Sahibzada Ajit Singh Nagar",
    "Sas Nagar Mohali":         "Sahibzada Ajit Singh Nagar",
    "Sas Nagar,Mohali":         "Sahibzada Ajit Singh Nagar",
    "Sas Nagar-Mohali":         "Sahibzada Ajit Singh Nagar",
    "Sas Nagarmohali":          "Sahibzada Ajit Singh Nagar",
    "Sasnagar Mohali":          "Sahibzada Ajit Singh Nagar",
    "Sasnagarmohali":           "Sahibzada Ajit Singh Nagar",
    "Mohali":                   "Sahibzada Ajit Singh Nagar",

    # Uttarakhand
    "Dehra Dun":                "Dehradun",

    # General spacing variants
    "Ram Nagar":                "Ramnagar",
    "Ram Garh":                 "Ramgarh",
    "Gandhi Nagar":             "Gandhinagar",
    "Ashok Nagar":              "Ashoknagar",
    "Jawahar Nagar":            "Jawaharnagar",
    "Naya Gaon":                "Nayagaon",
    "Kota 1":                   "Kota",
    "Sri Nagar":                "Srinagar",

    # Jharkhand
    "Raghunath Pur":            "Raghunathpur",
    "Saraikela-Kharsawan":      "Saraikela Kharsawan",
}

# ── Junk city names to delete entirely ────────────────────────────────────────
JUNK_CITIES = {
    "Urban", "Rural", "Metro", "Semi Urban", "Semiurban", "Semi-Urban",
    "Unknown", "Na", "N/A", "Nil", "None", "Test", "Dummy", "Other",
    "Metropolitan", "Suburban",
}

# ── Canonical city name mapping ──────────────────────────────────────────────
# Maps bad/variant names → correct official name (Title Case)
CITY_CANONICAL = {
    # Telangana / AP
    "Hyderabad-I":              "Hyderabad",
    "Hyderabad-Ii":             "Hyderabad",
    "Hyderabad - I":            "Hyderabad",
    "Hyderabad - Ii":           "Hyderabad",
    "Secunderabad":             "Secunderabad",
    "Secunderabad Cantonment":  "Secunderabad",
    "Kukatpally":               "Hyderabad",
    "Malkajgiri":               "Hyderabad",
    "Ameerpet":                 "Hyderabad",
    "Serilingampally":          "Hyderabad",
    "Seri Lingampally":         "Hyderabad",
    "Seri-Lingampally":         "Hyderabad",
    "Vizag":                    "Visakhapatnam",
    "Vishakhapatnam":           "Visakhapatnam",
    "Karim Nagar":              "Karimnagar",
    "Peddapalle":               "Peddapalli",
    "Peddapally":               "Peddapalli",
    "Ranga Reddy":              "Rangareddy",
    "Rangareddi":               "Rangareddy",
    "Mahabubnagar":             "Mahbubnagar",
    "Hanamkonda":               "Warangal",
    "Warangal Urban":           "Warangal",
    "Warangal Rural":           "Warangal",

    # Karnataka
    "Bangalore":                "Bengaluru",
    "Bangaluru":                "Bengaluru",
    "Bangalore Rural":          "Bengaluru",
    "Bangalore Urban":          "Bengaluru",
    "Mangalore":                "Mangaluru",
    "Mysore":                   "Mysuru",
    "Hubli":                    "Hubballi",
    "Belgaum":                  "Belagavi",
    "Bellary":                  "Ballari",
    "Gulbarga":                 "Kalaburagi",
    "Shimoga":                  "Shivamogga",
    "Tumkur":                   "Tumakuru",

    # Maharashtra
    "Mum Bai":                  "Mumbai",
    "Mumbai Suburban":          "Mumbai",
    "Mumbai City":              "Mumbai",
    "Bombay":                   "Mumbai",
    "Poona":                    "Pune",
    "Puna":                     "Pune",

    # Tamil Nadu
    "Madras":                   "Chennai",
    "Trichy":                   "Tiruchirappalli",
    "Trichirappalli":           "Tiruchirappalli",
    "Tanjore":                  "Thanjavur",
    "Pondicherry":              "Puducherry",

    # Kerala
    "Calicut":                  "Kozhikode",
    "Cochin":                   "Kochi",
    "Trivandrum":               "Thiruvananthapuram",
    "Trichur":                  "Thrissur",
    "Alleppey":                 "Alappuzha",
    "Quilon":                   "Kollam",

    # Delhi
    "New Delhi":                "Delhi",
    "Newdelhi":                 "Delhi",
    "North Delhi":              "Delhi",
    "South Delhi":              "Delhi",
    "East Delhi":               "Delhi",
    "West Delhi":               "Delhi",
    "Central Delhi":            "Delhi",
    "Shahdara":                 "Delhi",

    # Uttar Pradesh
    "Banaras":                  "Varanasi",
    "Benares":                  "Varanasi",
    "Allahabad":                "Prayagraj",

    # Odisha
    "Bhubaneshwar":             "Bhubaneswar",

    # Punjab
    "Mohali":                   "Sahibzada Ajit Singh Nagar",
    "S A S Nagar":              "Sahibzada Ajit Singh Nagar",

    # Gujarat
    "Baroda":                   "Vadodara",

    # West Bengal
    "Calcutta":                 "Kolkata",

    # General spacing variants
    "Ram Nagar":                "Ramnagar",
    "Gandhi Nagar":             "Gandhinagar",
    "Sri Nagar":                "Srinagar",
}

# ── Bank metadata from Excel spreadsheet ──────────────────────────────────────
# Primary source of truth: D:\IFSCCode\backup\Bank.xlsx
# Contains: bank_type, sub_type, website, logo_url, is_active, merged_into_id
# for all 1,352 banks — curated manually.

BANK_EXCEL_PATH = Path(__file__).parent.parent / "backup" / "Bank.xlsx"
# Fallback path if the backup folder is at a different location
BANK_EXCEL_FALLBACK = Path(r"D:\IFSCCode\backup\Bank.xlsx")


def load_bank_excel() -> dict:
    """Load bank metadata from the curated Excel spreadsheet.

    Returns dict keyed by bank id:
      { id: { name, short_name, bank_code, bank_type, sub_type,
              website, logo_url, is_active, merged_into_id, slug } }
    """
    xlsx_path = BANK_EXCEL_PATH if BANK_EXCEL_PATH.exists() else BANK_EXCEL_FALLBACK
    if not xlsx_path.exists():
        print(f"  {Fore.YELLOW}⚠ Bank.xlsx not found at {xlsx_path} — skipping Excel import{Style.RESET_ALL}")
        return {}

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    # Read header row to get column indices dynamically
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        bank_id = row[col["id"]]
        if bank_id is None:
            continue

        is_active_raw = row[col.get("is_active", -1)] if "is_active" in col else "true"
        is_active = str(is_active_raw).strip().lower() != "false"

        merged_into = row[col["merged_into_id"]] if "merged_into_id" in col else None
        if merged_into is not None:
            try:
                merged_into = int(merged_into)
            except (ValueError, TypeError):
                merged_into = None

        data[int(bank_id)] = {
            "name":           row[col["name"]] or "",
            "short_name":     row[col.get("short_name", -1)] if "short_name" in col else None,
            "bank_code":      row[col.get("bank_code", -1)] if "bank_code" in col else None,
            "bank_type":      row[col.get("bank_type", -1)] if "bank_type" in col else None,
            "sub_type":       row[col.get("sub_type", -1)] if "sub_type" in col else None,
            "website":        row[col.get("website", -1)] if "website" in col else None,
            "logo_url":       row[col.get("logo_url", -1)] if "logo_url" in col else None,
            "is_active":      is_active,
            "merged_into_id": merged_into,
            "slug":           row[col.get("slug", -1)] if "slug" in col else None,
        }

    wb.close()
    print(f"  Loaded {len(data):,} banks from {xlsx_path.name}")
    return data

# ── State ISO codes ───────────────────────────────────────────────────────────
STATE_CODES = {
    "Andhra Pradesh":           "IN-AP",
    "Arunachal Pradesh":        "IN-AR",
    "Assam":                    "IN-AS",
    "Bihar":                    "IN-BR",
    "Chhattisgarh":             "IN-CT",
    "Goa":                      "IN-GA",
    "Gujarat":                  "IN-GJ",
    "Haryana":                  "IN-HR",
    "Himachal Pradesh":         "IN-HP",
    "Jharkhand":                "IN-JH",
    "Karnataka":                "IN-KA",
    "Kerala":                   "IN-KL",
    "Madhya Pradesh":           "IN-MP",
    "Maharashtra":              "IN-MH",
    "Manipur":                  "IN-MN",
    "Meghalaya":                "IN-ML",
    "Mizoram":                  "IN-MZ",
    "Nagaland":                 "IN-NL",
    "Odisha":                   "IN-OR",
    "Punjab":                   "IN-PB",
    "Rajasthan":                "IN-RJ",
    "Sikkim":                   "IN-SK",
    "Tamil Nadu":               "IN-TN",
    "Telangana":                "IN-TG",
    "Tripura":                  "IN-TR",
    "Uttar Pradesh":            "IN-UP",
    "Uttarakhand":              "IN-UT",
    "West Bengal":              "IN-WB",
    # Union Territories
    "Andaman And Nicobar Islands": "IN-AN",
    "Andaman & Nicobar Islands":   "IN-AN",
    "Chandigarh":               "IN-CH",
    "Dadra And Nagar Haveli And Daman And Diu": "IN-DH",
    "Delhi":                    "IN-DL",
    "Jammu And Kashmir":        "IN-JK",
    "Jammu & Kashmir":          "IN-JK",
    "Ladakh":                   "IN-LA",
    "Lakshadweep":              "IN-LD",
    "Puducherry":               "IN-PY",
    "Pondicherry":              "IN-PY",
}


# ═════════════════════════════════════════════════════════════════════════════
# SLUG GENERATION
# ═════════════════════════════════════════════════════════════════════════════

def make_slug(name: str) -> str:
    """Convert a name to a URL-safe slug: lowercase, hyphens, no special chars."""
    s = name.strip().lower()
    s = re.sub(r'[^a-z0-9\s-]', '', s)   # remove special chars
    s = re.sub(r'[\s_]+', '-', s)          # spaces/underscores → hyphens
    s = re.sub(r'-{2,}', '-', s)           # collapse multiple hyphens
    return s.strip('-')


# ═════════════════════════════════════════════════════════════════════════════
# PINCODE EXTRACTION
# ═════════════════════════════════════════════════════════════════════════════

def extract_pincode(address: str) -> Optional[str]:
    """Extract 6-digit Indian pincode from address string."""
    if not address or len(address) < 6:
        return None

    # Pattern 1: explicit PIN/PINCODE label (handles spaces in number like "395 003")
    m = re.search(r'\bPIN(?:CODE)?[-:\s]+(\d[\s\d]{4,7}\d)', address, re.IGNORECASE)
    if m:
        digits = re.sub(r'\s+', '', m.group(1))
        if len(digits) == 6 and digits[0] != '0':
            return digits

    # Pattern 2: standalone 6-digit number (most common — appears at end)
    m = re.search(r'(?<!\d)([1-9]\d{5})(?!\d)', address)
    if m:
        return m.group(1)

    return None


# ═════════════════════════════════════════════════════════════════════════════
# DATABASE CONNECTION
# ═════════════════════════════════════════════════════════════════════════════

def get_conn():
    """Get a fresh direct PostgreSQL connection with retry for Neon cold starts."""
    url = os.environ.get("DIRECT_URL") or os.environ.get("DATABASE_URL", "")
    if not url:
        print(f"{Fore.RED}❌ No DATABASE_URL or DIRECT_URL found in environment.")
        sys.exit(1)
    # Strip pgbouncer param — psycopg2 doesn't understand it
    url = url.replace("&pgbouncer=true", "").replace("?pgbouncer=true", "")

    for attempt in range(3):
        try:
            conn = psycopg2.connect(url, connect_timeout=30,
                                    keepalives=1,
                                    keepalives_idle=30,
                                    keepalives_interval=10,
                                    keepalives_count=5)
            conn.autocommit = False
            # Warm up connection — forces Neon to fully wake
            cur = conn.cursor()
            cur.execute("SELECT 1")
            cur.close()
            return conn
        except psycopg2.OperationalError as e:
            if attempt < 2:
                print(f"  {Fore.YELLOW}⚠ Connection attempt {attempt+1} failed, retrying in 3s...{Style.RESET_ALL}")
                time.sleep(3)
            else:
                raise


def ensure_conn(conn):
    """Check if connection is alive, reconnect if dead (Neon idle timeout)."""
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1")
        cur.close()
        return conn
    except (psycopg2.OperationalError, psycopg2.InterfaceError):
        print(f"  {Fore.YELLOW}⚠ Connection dropped, reconnecting...{Style.RESET_ALL}")
        try:
            conn.close()
        except Exception:
            pass
        return get_conn()


# ═════════════════════════════════════════════════════════════════════════════
# STEP 1 — FIX DISTRICTS
# ═════════════════════════════════════════════════════════════════════════════

def fix_districts(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Step 1 — Fixing districts")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()

    # 1a. Count before
    cur.execute('SELECT COUNT(*) FROM "District"')
    before = cur.fetchone()[0]
    print(f"  Districts before: {before:,}")

    # 1b. Delete junk districts — reassign branches to NULL first
    junk_deleted = 0
    for junk_name in JUNK_DISTRICTS:
        cur.execute(
            'SELECT id FROM "District" WHERE lower(name) = lower(%s)',
            (junk_name,)
        )
        rows = cur.fetchall()
        for (did,) in rows:
            cur.execute(
                'UPDATE "Branch" SET district_id = NULL WHERE district_id = %s',
                (did,)
            )
            cur.execute('DELETE FROM "District" WHERE id = %s', (did,))
            junk_deleted += 1

    conn.commit()
    print(f"  Junk districts removed: {junk_deleted}")

    # 1c. Apply canonical name mapping
    merged = 0
    for bad_name, canonical_name in DISTRICT_CANONICAL.items():
        # Find all districts with the bad name
        cur.execute(
            'SELECT id, state_id FROM "District" WHERE name = %s',
            (bad_name,)
        )
        bad_rows = cur.fetchall()
        if not bad_rows:
            continue

        for (bad_id, state_id) in bad_rows:
            # Find or create the canonical district in the same state
            cur.execute(
                'SELECT id FROM "District" WHERE name = %s AND state_id = %s',
                (canonical_name, state_id)
            )
            canonical_row = cur.fetchone()

            if canonical_row:
                canonical_id = canonical_row[0]
                # Reassign all branches from bad district to canonical
                cur.execute(
                    'UPDATE "Branch" SET district_id = %s WHERE district_id = %s',
                    (canonical_id, bad_id)
                )
                cur.execute('DELETE FROM "District" WHERE id = %s', (bad_id,))
            else:
                # Rename the bad district to the canonical name
                cur.execute(
                    'UPDATE "District" SET name = %s WHERE id = %s',
                    (canonical_name, bad_id)
                )
            merged += 1

    conn.commit()
    print(f"  District names normalised: {merged}")

    # 1d. Merge duplicates — same name+state_id, keep lowest id
    cur.execute("""
        SELECT name, state_id, array_agg(id ORDER BY id) as ids
        FROM "District"
        GROUP BY name, state_id
        HAVING COUNT(*) > 1
    """)
    dupes = cur.fetchall()
    deduped = 0
    for (name, state_id, ids) in dupes:
        keep_id = ids[0]
        for dup_id in ids[1:]:
            cur.execute(
                'UPDATE "Branch" SET district_id = %s WHERE district_id = %s',
                (keep_id, dup_id)
            )
            cur.execute('DELETE FROM "District" WHERE id = %s', (dup_id,))
            deduped += 1

    conn.commit()
    print(f"  Duplicate districts merged: {deduped}")

    # 1e. Populate slug + normalized_name for districts (add columns if missing)
    try:
        cur.execute('ALTER TABLE "District" ADD COLUMN IF NOT EXISTS slug VARCHAR(120)')
        cur.execute('ALTER TABLE "District" ADD COLUMN IF NOT EXISTS normalized_name VARCHAR(100)')
        conn.commit()
    except Exception:
        conn.rollback()

    # Bulk update: generate slug + normalized_name server-side in one statement
    cur.execute("""
        UPDATE "District"
        SET normalized_name = UPPER(TRIM(name)),
            slug = LOWER(TRIM(
                REGEXP_REPLACE(
                    REGEXP_REPLACE(
                        REGEXP_REPLACE(TRIM(name), '[^a-zA-Z0-9\\s-]', '', 'g'),
                        '[\\s_]+', '-', 'g'),
                    '-{2,}', '-', 'g')
            ))
    """)
    slug_updates = cur.rowcount
    conn.commit()
    print(f"  District slugs populated: {slug_updates}")

    # 1f. Count after
    cur.execute('SELECT COUNT(*) FROM "District"')
    after = cur.fetchone()[0]
    print(f"  Districts after:  {after:,}  (removed {before - after:,})")
    cur.close()


# ═════════════════════════════════════════════════════════════════════════════
# STEP 1b — FIX CITIES
# ═════════════════════════════════════════════════════════════════════════════

def fix_cities(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Step 1b — Fixing cities")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()

    # Count before
    cur.execute('SELECT COUNT(*) FROM "City"')
    before = cur.fetchone()[0]
    print(f"  Cities before: {before:,}")

    # 1. Delete junk cities — reassign branches to NULL first
    junk_deleted = 0
    for junk_name in JUNK_CITIES:
        cur.execute(
            'SELECT id FROM "City" WHERE lower(name) = lower(%s)',
            (junk_name,)
        )
        rows = cur.fetchall()
        for (cid,) in rows:
            cur.execute(
                'UPDATE "Branch" SET city_id = NULL WHERE city_id = %s',
                (cid,)
            )
            cur.execute('DELETE FROM "City" WHERE id = %s', (cid,))
            junk_deleted += 1

    conn.commit()
    print(f"  Junk cities removed: {junk_deleted}")

    # 2. Apply canonical name mapping
    merged = 0
    for bad_name, canonical_name in CITY_CANONICAL.items():
        cur.execute(
            'SELECT id, state_id FROM "City" WHERE name = %s',
            (bad_name,)
        )
        bad_rows = cur.fetchall()
        if not bad_rows:
            continue

        for (bad_id, state_id) in bad_rows:
            cur.execute(
                'SELECT id FROM "City" WHERE name = %s AND state_id = %s',
                (canonical_name, state_id)
            )
            canonical_row = cur.fetchone()

            if canonical_row:
                canonical_id = canonical_row[0]
                cur.execute(
                    'UPDATE "Branch" SET city_id = %s WHERE city_id = %s',
                    (canonical_id, bad_id)
                )
                cur.execute('DELETE FROM "City" WHERE id = %s', (bad_id,))
            else:
                cur.execute(
                    'UPDATE "City" SET name = %s WHERE id = %s',
                    (canonical_name, bad_id)
                )
            merged += 1

    conn.commit()
    print(f"  City names normalised: {merged}")

    # 3. Merge duplicates — same name+state_id, keep lowest id
    cur.execute("""
        SELECT name, state_id, array_agg(id ORDER BY id) as ids
        FROM "City"
        GROUP BY name, state_id
        HAVING COUNT(*) > 1
    """)
    dupes = cur.fetchall()
    deduped = 0
    for (name, state_id, ids) in dupes:
        keep_id = ids[0]
        for dup_id in ids[1:]:
            cur.execute(
                'UPDATE "Branch" SET city_id = %s WHERE city_id = %s',
                (keep_id, dup_id)
            )
            cur.execute('DELETE FROM "City" WHERE id = %s', (dup_id,))
            deduped += 1

    conn.commit()
    print(f"  Duplicate cities merged: {deduped}")

    # 4. Populate slug + normalized_name for all cities (bulk server-side)
    cur.execute("""
        UPDATE "City"
        SET normalized_name = UPPER(TRIM(name)),
            slug = LOWER(TRIM(
                REGEXP_REPLACE(
                    REGEXP_REPLACE(
                        REGEXP_REPLACE(TRIM(name), '[^a-zA-Z0-9\\s-]', '', 'g'),
                        '[\\s_]+', '-', 'g'),
                    '-{2,}', '-', 'g')
            ))
    """)
    slug_updates = cur.rowcount
    conn.commit()
    print(f"  City slugs populated: {slug_updates}")

    # Count after
    cur.execute('SELECT COUNT(*) FROM "City"')
    after = cur.fetchone()[0]
    print(f"  Cities after:  {after:,}  (removed {before - after:,})")
    cur.close()


# ═════════════════════════════════════════════════════════════════════════════
# STEP 2 — EXTRACT PINCODES FROM ADDRESS
# ═════════════════════════════════════════════════════════════════════════════

def fix_pincodes(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Step 2 — Extracting pincodes from addresses")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()

    # Fetch all branches with address but no pincode
    cur.execute("""
        SELECT id, address FROM "Branch"
        WHERE pincode IS NULL
        AND address IS NOT NULL
        AND address != ''
    """)
    rows = cur.fetchall()
    print(f"  Branches needing pincode extraction: {len(rows):,}")

    updates = []
    found = 0
    for (branch_id, address) in rows:
        pincode = extract_pincode(address)
        if pincode:
            updates.append((pincode, branch_id))
            found += 1

    if updates:
        psycopg2.extras.execute_batch(
            cur,
            'UPDATE "Branch" SET pincode = %s WHERE id = %s',
            updates,
            page_size=1000
        )
        conn.commit()

    print(f"  Pincodes extracted and saved: {found:,}")
    print(f"  Branches still without pincode: {len(rows) - found:,} (genuinely missing in source data)")
    cur.close()


# ═════════════════════════════════════════════════════════════════════════════
# STEP 3 — ADD BANK METADATA
# ═════════════════════════════════════════════════════════════════════════════

def fix_bank_metadata(conn):
    """Populate bank metadata from Bank.xlsx — the curated spreadsheet.

    Updates: bank_type, sub_type, short_name, website, logo_url,
             is_active, merged_into_id for ALL banks matched by id.

    Merged banks: branches from inactive merged banks get reassigned
    to the target bank, then the inactive bank is soft-deactivated.
    """
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Step 3 — Applying bank metadata from Bank.xlsx")
    print(f"{'═'*60}{Style.RESET_ALL}")

    excel_data = load_bank_excel()
    if not excel_data:
        print(f"  {Fore.YELLOW}⚠ No Excel data — skipping bank metadata step{Style.RESET_ALL}")
        return

    cur = conn.cursor()
    cur.execute('SELECT id, name FROM "Bank"')
    db_banks = cur.fetchall()

    updated = 0
    deactivated = 0
    merged = 0
    type_counts = {}

    for (bank_id, bank_name) in db_banks:
        meta = excel_data.get(bank_id)
        if not meta:
            continue

        bank_type = meta["bank_type"] or None
        sub_type  = meta["sub_type"] or None
        short_name = meta["short_name"] or None
        website   = meta["website"] or None
        logo_url  = meta["logo_url"] or None
        is_active = meta["is_active"]
        merged_into = meta["merged_into_id"]

        if bank_type:
            type_counts[bank_type] = type_counts.get(bank_type, 0) + 1

        # Update metadata fields
        cur.execute("""
            UPDATE "Bank"
            SET bank_type      = COALESCE(%s, bank_type),
                sub_type       = COALESCE(%s, sub_type),
                short_name     = COALESCE(%s, short_name),
                website        = COALESCE(%s, website),
                logo_url       = COALESCE(%s, logo_url),
                is_active      = %s,
                merged_into_id = %s
            WHERE id = %s
        """, (bank_type, sub_type, short_name, website, logo_url,
              is_active, merged_into, bank_id))
        updated += 1

        # Handle merged banks: reassign branches to the target bank
        if not is_active and merged_into:
            # Check target bank exists
            cur.execute('SELECT id FROM "Bank" WHERE id = %s', (merged_into,))
            if cur.fetchone():
                cur.execute(
                    'UPDATE "Branch" SET bank_id = %s WHERE bank_id = %s',
                    (merged_into, bank_id)
                )
                reassigned = cur.rowcount
                cur.execute(
                    'DELETE FROM "bank_state_presence" WHERE bank_id = %s',
                    (bank_id,)
                )
                merged += 1
                if reassigned > 0:
                    print(f"    Merged: {bank_name} -> id={merged_into} ({reassigned} branches moved)")

        if not is_active:
            deactivated += 1

    conn.commit()

    # Stats
    print(f"\n  Banks matched & updated: {updated:,}")
    print(f"  Banks deactivated:       {deactivated}")
    print(f"  Merged bank reassignments: {merged}")
    print(f"\n  Type distribution:")
    for btype, count in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"    {btype:<25} {Fore.GREEN}{count:>6,}{Style.RESET_ALL}")

    cur.execute('SELECT COUNT(*) FROM "Bank" WHERE website IS NOT NULL AND website != %s', ('',))
    with_website = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Bank" WHERE logo_url IS NOT NULL AND logo_url != %s', ('',))
    with_logo = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Bank" WHERE is_active = true')
    active = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "Bank"')
    total = cur.fetchone()[0]
    print(f"\n  Banks with website:  {with_website}/{total}")
    print(f"  Banks with logo:     {with_logo}/{total}")
    print(f"  Active banks:        {active}/{total}")
    cur.close()


# ═════════════════════════════════════════════════════════════════════════════
# STEP 4 — ADD STATE ISO CODES
# ═════════════════════════════════════════════════════════════════════════════

def fix_state_codes(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Step 4 — Adding state ISO codes")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()
    cur.execute('SELECT id, name FROM "State"')
    states = cur.fetchall()

    updated = 0
    not_found = []
    for (state_id, state_name) in states:
        code = STATE_CODES.get(state_name)
        if code:
            cur.execute(
                'UPDATE "State" SET code = %s WHERE id = %s',
                (code, state_id)
            )
            updated += 1
        else:
            not_found.append(state_name)

    conn.commit()
    print(f"  States updated with ISO codes: {updated}")
    if not_found:
        print(f"  States without code mapping: {not_found}")
    cur.close()



def remove_defunct_banks(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Removing defunct/closed banks")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()
    cur.execute('SELECT id, name FROM "Bank" WHERE name = ANY(%s)', (list(DEFUNCT_BANKS),))
    defunct = cur.fetchall()

    if not defunct:
        print(f"  No defunct banks found")
        cur.close()
        return

    for bank_id, name in defunct:
        cur.execute('DELETE FROM "bank_state_presence" WHERE bank_id = %s', (bank_id,))
        cur.execute('DELETE FROM "Branch" WHERE bank_id = %s', (bank_id,))
        branches_deleted = cur.rowcount
        cur.execute('DELETE FROM "Bank" WHERE id = %s', (bank_id,))
        print(f"  ✓ {name} ({branches_deleted} branches removed)")

    conn.commit()
    cur.close()

# ═════════════════════════════════════════════════════════════════════════════
# STEP 5 — PRINT FINAL STATS
# ═════════════════════════════════════════════════════════════════════════════

def print_final_stats(conn):
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Final Database Stats")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()
    stats = [
        ("Total branches",             'SELECT COUNT(*) FROM "Branch"'),
        ("Branches with district",     'SELECT COUNT(*) FROM "Branch" WHERE district_id IS NOT NULL'),
        ("Branches with pincode",      'SELECT COUNT(*) FROM "Branch" WHERE pincode IS NOT NULL'),
        ("Branches with phone",        'SELECT COUNT(*) FROM "Branch" WHERE phone IS NOT NULL'),
        ("Total districts (clean)",    'SELECT COUNT(*) FROM "District"'),
        ("Total cities (clean)",       'SELECT COUNT(*) FROM "City"'),
        ("Total banks",                'SELECT COUNT(*) FROM "Bank"'),
        ("Active banks",               'SELECT COUNT(*) FROM "Bank" WHERE is_active = true'),
        ("Inactive/merged banks",      'SELECT COUNT(*) FROM "Bank" WHERE is_active = false'),
        ("Banks with website",         'SELECT COUNT(*) FROM "Bank" WHERE website IS NOT NULL'),
        ("Banks with logo",            'SELECT COUNT(*) FROM "Bank" WHERE logo_url IS NOT NULL'),
        ("States with ISO code",       'SELECT COUNT(*) FROM "State" WHERE code IS NOT NULL'),
    ]

    for label, sql in stats:
        cur.execute(sql)
        val = cur.fetchone()[0]
        print(f"  {label:<35} {Fore.GREEN}{val:>10,}{Style.RESET_ALL}")

    cur.close()

def fix_corrupted_bank_codes(conn):
    """Remove or null out bank_code for cooperative/small banks with corrupted codes"""
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  Fixing corrupted bank codes")
    print(f"{'═'*60}{Style.RESET_ALL}")

    cur = conn.cursor()

    # Find cooperative banks
    cur.execute('''
        SELECT b.id, b.name, b.bank_code, COUNT(br.id) as branch_count
        FROM "Bank" b
        LEFT JOIN "Branch" br ON b.id = br.bank_id
        WHERE b.name LIKE '%Co-operative%' 
           OR b.name LIKE '%co-op%'
           OR b.name LIKE '%Urban Bank%'
        GROUP BY b.id, b.name, b.bank_code
        ORDER BY branch_count ASC
    ''')
    
    coop_banks = cur.fetchall()
    print(f"  Found {len(coop_banks)} cooperative/small banks")

    deleted = 0
    nulled = 0

    for bank_id, name, code, branch_count in coop_banks:
        # If < 5 branches, delete it
        if branch_count < 5:
            cur.execute('DELETE FROM "bank_state_presence" WHERE bank_id = %s', (bank_id,))
            cur.execute('DELETE FROM "Branch" WHERE bank_id = %s', (bank_id,))
            cur.execute('DELETE FROM "Bank" WHERE id = %s', (bank_id,))
            deleted += 1
            print(f"  ✓ Deleted: {name} ({branch_count} branches)")
        else:
            # Otherwise set bank_code to NULL so it doesn't get wrong metadata
            cur.execute('UPDATE "Bank" SET bank_code = NULL WHERE id = %s', (bank_id,))
            nulled += 1
            print(f"  ✓ Nulled code: {name} ({branch_count} branches)")

    conn.commit()
    print(f"  Deleted: {deleted} banks, Nulled codes: {nulled} banks")
    cur.close()

# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    print(f"\n{Fore.CYAN}{'═'*60}")
    print(f"  BankInfoHub — Data Quality Cleanup")
    print(f"{'═'*60}{Style.RESET_ALL}")

    conn = get_conn()
    t0 = time.time()

    try:
        fix_districts(conn)
        fix_cities(conn)
        fix_bank_metadata(conn)      # Excel-driven: updates metadata + handles merges + deactivates
        remove_defunct_banks(conn)    # Fallback: deletes any remaining defunct banks by name
        fix_pincodes(conn)
        fix_state_codes(conn)
        print_final_stats(conn)

        duration = time.time() - t0
        print(f"\n{Fore.GREEN}✅ Cleanup complete in {duration:.1f}s{Style.RESET_ALL}\n")

        # Clear API cache so changes are visible immediately
        api_url = os.environ.get("API_URL", "http://localhost:3001")
        try:
            import urllib.request
            req = urllib.request.Request(f"{api_url}/api/cache/clear", method="POST")
            urllib.request.urlopen(req, timeout=5)
            print(f"✅ API cache cleared — changes visible immediately")
        except Exception:
            print(f"ℹ️  API cache not cleared (server not reachable) — restart backend to see changes")

    except Exception as e:
        conn.rollback()
        print(f"\n{Fore.RED}❌ Cleanup failed: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
